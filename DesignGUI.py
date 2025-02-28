# -*- coding: utf-8 -*-
"""
Created on Wed May 19 23:25:42 2021

@author: PRASHIK
"""
import math as m

from tkinter import *
#from tkinter import IntVar
import openpyxl as op

window=Tk()
window.title("Rigid Flange Coupling Designer")
window.geometry("500x500")
window.iconbitmap(r'CTR.ico')

l1=Label(window,text="Rigid Flange Coupling",font=("Arial",12))
l1.grid(row=1,column=1,sticky=W)

l2=Label(window,text="Define Part Number",font=("Arial",10))
l2.grid(row=2,column=1,sticky=W,pady=2)

l3=Label(window,text="Enter Power",font=("Arial",10))
l3.grid(row=3,column=1,sticky=W,pady=2)

l4=Label(window,text="Enter RPM",font=("Arial",10))
l4.grid(row=4,column=1,sticky=W,pady=2)

l5=Label(window,text="Enter Service Factor",font=("Arial",10))
l5.grid(row=5,column=1,sticky=W,pady=2)

l6=Label(window,text="Enter Sigma for Shaft",font=("Arial",10))
l6.grid(row=6,column=1,sticky=W,pady=2)

l7=Label(window,text="Enter Sigma for Key",font=("Arial",10))
l7.grid(row=7,column=1,sticky=W,pady=2)

l8=Label(window,text="Enter Sigma for Cast",font=("Arial",10))
l8.grid(row=8,column=1,sticky=W,pady=2)

l8=Label(window,text="Enter Sigma for Cast",font=("Arial",10))
l8.grid(row=8,column=1,sticky=W,pady=2)

l9=Label(window,text="Enter Number of Bolts",font=("Arial",10))
l9.grid(row=9,column=1,sticky=W,pady=2)

l10=Label(window,text="Powered By TruVisiory Projections",fg="Red",font=("Arial",10))
l10.grid(row=40,column=3,sticky=E,pady=10)

k=1
name=StringVar()
P=IntVar()
n=IntVar()
K=IntVar()
MTRL_shaft=IntVar()
MTRL_key=IntVar()
MTRL_hub=IntVar()
N=IntVar()

e2=Entry(window,textvariable=name,font=("Arial",10))
e2.grid(row=2,column=2,sticky=W,pady=2)

e3=Entry(window,textvariable=P,font=("Arial",10))
e3.grid(row=3,column=2,sticky=W,pady=2)

e4=Entry(window,textvariable=n,font=("Arial",10))
e4.grid(row=4,column=2,sticky=W,pady=2)

e5=Entry(window,textvariable=K,font=("Arial",10))
e5.grid(row=5,column=2,sticky=W,pady=2)

e6=Entry(window,textvariable=MTRL_shaft,font=("Arial",10))
e6.grid(row=6,column=2,sticky=W,pady=2)

e7=Entry(window,textvariable=MTRL_key,font=("Arial",10))
e7.grid(row=7,column=2,sticky=W,pady=2)

e8=Entry(window,textvariable=MTRL_hub,font=("Arial",10))
e8.grid(row=8,column=2,sticky=W,pady=2)

e8=Entry(window,textvariable=N,font=("Arial",10))
e8.grid(row=9,column=2,sticky=W,pady=2)


def onClick():
    #welcome to CI Coupling Autodesigner Module by Shreyas Koshti, TRUVISIORY PROJECTIONS.
    #Defination of Parameters: units defines "Input/Calculated" Unit conversion in inscribed into calculation
    #Input Terms:        P=Power Input Unit KW/W
    #Service factor:     K= No units
    #RPM of Shaft  :     n=rpm/rpm
    #Torqe         :     T=N/mmSq
    
    # Section 1: RECEIVING INPUT DATA,
    # SECTION 1.1 : LEVEL 1 CALCULATION FOR DETERMINING REQUIRED TERMS.
    # SHEAR STRESS AND FOS:  ACCORDING TO SHEAR STRESS THEORY
    SHR_STR=MTRL_shaft.get()/2
    #CALCULATION OF TORQUE AND SERVICE FACTOR:
    T=((P.get()*60000000)/(2*m.pi*(n.get())))
    Tmax=T*K.get()
    #SECTION 2: CALCULATING DIAMETER OF THE SHAFT:                   Load EXCEL D=VALUE
    d=((16*Tmax)/(m.pi*SHR_STR))**(1/3)
    print(m.ceil(d))
    #SECTION 3: DESIGN OF SQUARE KEY                                 Load EXCEL Key=VALUE
    w=h=d/4
    len_hub=1.5*w
    #CHECKING KEY FOR CRUSHING STRESS AND UPGRADING VALUES.
    cru_str=MTRL_key.get()
    
    len_hub1=((4*Tmax)/(d*h*cru_str))
    if len_hub1>len_hub:
        len_hub=len_hub1
    else:
        len_hub1=len_hub
    #CHECKING KEY FOR SHARING STRESS
    len_hub2=len_hub1 = (2*Tmax)/(d*h*SHR_STR)
    if len_hub1>len_hub:
        len_hub=len_hub2
    else:
        len_hub2=len_hub
    w=h = m.ceil(w)
    
    len_hub=m.ceil(len_hub)
   
    #LOAD VALUE OF DIAMETER OF HUB                                  LOAD TO EXCEL.
    #SECTION 5: DESIGN OF FLANGE
    d_hub=2*d
    
    tf=0.5*d             #THICKNESS OF FLANGE.
    tp=0.25*d            #PROTECTIVE THICKNESS OF COLLER.
    d1=3*d               #BOLT PCD IN MM
    d2=4*d               #OUTER DIAMETER OF FLANGE
    d3=1.1*d_hub         #DIAMETER OF FLANGE RECESS OR PILOT DIAMETER
    id1=d/10             #FLANGE ALLINMENT INDENTTION
    id2=(d/10)-1.5       #FLANGE NEGATIVE PROTRUSION
    
    #SECTION 4: DESIGNING COUPLING HUB
    #ALREADY CALCULATED LENGTH STORED IN len_hub
    
    #CHECKING FOR SHEAR STRESS FAILURE OF HUB
    SHR_STR_hub=((16*Tmax)/(m.pi*((d_hub)**3)*(1-(0.0625))))
    if SHR_STR_hub<(MTRL_hub.get()/2):
        tf = tf
    else:
        d_hub=d_hub+5
    
    
    #CHECKING FLANGE FOR SHEAR FAILURE
    SHR_STR_hub1=((Tmax*2)/(m.pi*d_hub*d_hub*tf))
    if SHR_STR_hub1<(MTRL_hub.get()/2):
        tf=tf
        
    else:
        tf=tf+3
    #SECTION 6: DESIGN OF BOLTS.
    boltcat=[4,5,6,8,10,12,16,20,22,24,25,26,28,30,32,34,35,36]
    #CHECKING FOR SHEAR FAILURE
    db=((Tmax*8)/((N.get())*m.pi*d1*SHR_STR))
    db=m.sqrt(db)
    db=m.ceil(db)+1
    cat1=4
    cat=boltcat[cat1]
    for cat in boltcat:
      if cat < db:
        cat1+=1
      else:
          break
        #CHECKING FOR CRUSHING FAILURE
    MTRL_shaft1=((2*Tmax)/((N.get())*d1*db*tf))
    if MTRL_shaft1<MTRL_shaft.get():
        db=db
    else:
        db+=1
    
    def reve(f):
        return m.ceil(f / 2.) * 2
    
    #SECTION 7: working with excel file
    c = op.load_workbook('C:\\Users\\Shreyas\\Desktop\\ProgramAssignment\\RigidFlangeCoupling.xlsx')
    s = c['Sheet1']
    s.cell(row=17, column=4).value = reve(d)
    s.cell(row=18, column=4).value = reve(d_hub)
    s.cell(row=19, column=4).value = m.ceil(len_hub)
    s.cell(row=20, column=4).value = 'M%d' % cat
    s.cell(row=21, column=4).value = m.ceil(tp)
    s.cell(row=22, column=4).value = m.ceil(tf)
    s.cell(row=23, column=4).value = reve(d1)
    s.cell(row=24, column=4).value = reve(d2)
    s.cell(row=25, column=4).value = reve(w)
    s.cell(row=26, column=4).value = reve(len_hub+5)
    s.cell(row=27, column=4).value = reve(d3)
    s.cell(row=28, column=4).value = m.ceil(id1)
    s.cell(row=29, column=4).value = m.ceil(id2)
    s.cell(row=4,column=4).value=P.get()
    s.cell(row=5,column=4).value=n.get()
    s.cell(row=6,column=4).value=K.get()
    s.cell(row=1, column=4).value = name.get()
    c.save('C:\\Users\\Shreyas\\Desktop\\ProgramAssignment\\RigidFlangeCoupling.xlsx')
    print("Program Initiated")
    
  
          
b1=Button(window,text="Calculate and Execute", command= onClick ,font=("Arial",10),fg="white",bg="Gray")
b1.grid(row=10,column=2,sticky=W,pady=4)



window.mainloop()