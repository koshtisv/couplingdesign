#welcome to CI Coupling Autodesigner Module by Shreyas Koshti, TRUVISIORY PROJECTIONS.
#Defination of Parameters: units defines "Input/Calculated" Unit conversion in inscribed into calculation
#Input Terms:        P=Power Input Unit KW/W
#Service factor:     K= No units
#RPM of Shaft  :     n=rpm/rpm
#Torqe         :     T=N/mmSq
import math as m
# Section 1: RECEIVING INPUT DATA,
k=1
name=input(print("ASSIGN PART NUMBER TO THIS MODEL :"))
P=int(input(print("ENTER THE VALUE OF POWER IN KW : USE KW=HP x 1.34:  ")))
n=int(input(print("FEED SHAFT RPM:")))
K=float(input(print("FEED IN SERVICE FACTOR FOR THE COUPLING(DEFAULT VALUE=1)::")))
MTRL_shaft=int(input(print("ENTER VALUE OF PASS STRESS FOR SHAFT VALUE IN N/SQ.MM(MPA)")))
MTRL_key=int(input(print("ENTER VALUE OF PASS STRESS FOR KEY VALUE IN N/SQ.MM(MPA)")))
MTRL_hub=int(input(print("ENTER VALUE OF PASS STRESS FOR CI FLANGE VALUE IN N/SQ.MM(MPA)")))
N=int(input(print('HOW MANY NO OF BOLTS YOU NEED :')))
# SECTION 1.1 : LEVEL 1 CALCULATION FOR DETERMINING REQUIRED TERMS.
# SHEAR STRESS AND FOS:  ACCORDING TO SHEAR STRESS THEORY
SHR_STR=MTRL_shaft/2
#CALCULATION OF TORQUE AND SERVICE FACTOR:
T=((P*60000000)/(2*m.pi*n))
Tmax=T*K
#SECTION 2: CALCULATING DIAMETER OF THE SHAFT:                   Load EXCEL D=VALUE
d=((16*Tmax)/(m.pi*SHR_STR))**(1/3)
print(m.ceil(d))
#SECTION 3: DESIGN OF SQUARE KEY                                 Load EXCEL Key=VALUE
w=h=d/4
len_hub=1.5*w
#CHECKING KEY FOR CRUSHING STRESS AND UPGRADING VALUES.
cru_str=MTRL_key

len_hub1=((4*Tmax)/(d*h*cru_str))
if len_hub1>len_hub:
    len_hub=len_hub1
else:
    len_hub1=len_hub
#CHECKING KEY FOR SHARING STRESS
len_hub2=len_hub1 = (2*Tmax)/(d*h*SHR_STR)
if len_hub1>len_hub:
    len_hub=len_hub2
else:
    len_hub2=len_hub
w=h = m.ceil(w)

len_hub=m.ceil(len_hub)
#SECTION 4: DESIGNING COUPLING HUB
#ALREADY CALCULATED LENGTH STORED IN len_hub
d_hub=2*d
#CHECKING FOR SHEAR STRESS FAILURE OF HUB
SHR_STR_hub=((16*Tmax)/(m.pi*((d_hub)**3)*(1-(0.0625))))
if SHR_STR_hub<(MTRL_hub/2):
    print("HUB DIAMETER ASSUMED, SEEMS SAFE")
else:
    print("HUB DIAMETER MAY FAIL, RE-EVALUATING ")
    d_hub=d_hub+5
#LOAD VALUE OF DIAMETER OF HUB                                  LOAD TO EXCEL.
#SECTION 5: DESIGN OF FLANGE
tf=0.5*d             #THICKNESS OF FLANGE.
tp=0.25*d            #PROTECTIVE THICKNESS OF COLLER.
d1=3*d               #BOLT PCD IN MM
d2=4*d               #OUTER DIAMETER OF FLANGE
d3=1.1*d_hub         #DIAMETER OF FLANGE RECESS OR PILOT DIAMETER
id1=d/10             #FLANGE ALLINMENT INDENTTION
id2=(d/10)-1.5       #FLANGE NEGATIVE PROTRUSION
#CHECKING FLANGE FOR SHEAR FAILURE
SHR_STR_hub1=((Tmax*2)/(m.pi*d_hub*d_hub*tf))
if SHR_STR_hub1<(MTRL_hub/2):
    print("FLANGE THICKNESS ASSUMED, SEEMS SAFE")
else:
    print("FLANGE THICKNESS MAY FAIL, RE-EVALUATING ")
    tf=tf+3
#SECTION 6: DESIGN OF BOLTS.
boltcat=[4,5,6,8,10,12,16,20,22,24,25,26,28,30,32,34,35,36]
#CHECKING FOR SHEAR FAILURE
db=((Tmax*8)/(N*m.pi*d1*SHR_STR))
db=m.sqrt(db)
db=m.ceil(db)+1
cat1=4
cat=boltcat[cat1]
for cat in boltcat:
  if cat < db:
    cat1+=1
  else:
      break
print("SELECTED BOLT SIZE IS M%d"%cat)
#CHECKING FOR CRUSHING FAILURE
MTRL_shaft1=((2*Tmax)/(N*d1*db*tf))
if MTRL_shaft1<MTRL_shaft:
    db=db
else:
    db+=1

#BOLTS ARE NOW SELECTED FOR THE ASSEMBLY AND FINAL ROUND UP OF VALUES               PASS TO EXCEL
print("Shaft Diameter Calculated = %d"%d)
print("Hub Diameter= %d"%d_hub)
print("Bolt Hole PCD= %d"%d1)
print("Outer Diameter of flange= %d"%d2)
print("Outer Diameter of Pilot= %d"%d3)
print("Flange thickness= %d"%tf)
print("Coller Thickness= %d"%tp)
permission=input("Do You want to write data to excel. Press Y for Yes and N for No")
if permission == 'Y' or 'N':
    if permission == 'Y' :

        def reve(f):
            return m.ceil(f / 2.) * 2

        import openpyxl as op
        c = op.load_workbook('C:\\Users\\Shreyas\\Desktop\\ProgramAssignment\\RigidFlangeCoupling.xlsx')
        s = c['Sheet1']
        s.cell(row=17, column=4).value = reve(d)
        s.cell(row=18, column=4).value = reve(d_hub)
        s.cell(row=19, column=4).value = m.ceil(len_hub)
        s.cell(row=20, column=4).value = 'M%d' % cat
        s.cell(row=21, column=4).value = m.ceil(tp)
        s.cell(row=22, column=4).value = m.ceil(tf)
        s.cell(row=23, column=4).value = reve(d1)
        s.cell(row=24, column=4).value = reve(d2)
        s.cell(row=25, column=4).value = reve(w)
        s.cell(row=26, column=4).value = reve(len_hub+5)
        s.cell(row=27, column=4).value = reve(d3)
        s.cell(row=28, column=4).value = m.ceil(id1)
        s.cell(row=29, column=4).value = m.ceil(id2)
        s.cell(row=4,column=4).value=P
        s.cell(row=5,column=4).value=n
        s.cell(row=6,column=4).value=K
        s.cell(row=1, column=4).value = name
        c.save('C:\\Users\\Shreyas\\Desktop\\ProgramAssignment\\RigidFlangeCoupling.xlsx')
        print("Program Initiated")

        print("Successfully written")
    elif permission=='N':
        print("Sure, Write Cancelled")
else:
    print("Choice Invalid")


